import random
import hashlib
from django.db import transaction, IntegrityError
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView

from authentication.permissions import IsAdmin, IsVoter
from authentication.utils import log_event
from authentication.models import AuditLog
from voters.models import Voter, Constituency
from elections.models import Election, Candidate, VoteReceipt, Vote
from elections.serializers import ElectionSerializer, CandidateSerializer, CandidateCreateSerializer

class ElectionListCreateView(ListCreateAPIView):
    queryset = Election.objects.all().order_by('-created_at')
    serializer_class = ElectionSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.IsAuthenticated()]
        return [IsAdmin()]


class ElectionDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Election.objects.all()
    serializer_class = ElectionSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.IsAuthenticated()]
        return [IsAdmin()]


class VoterElectionsListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user = request.user
        
        # Admin gets all elections directly
        if user.role == 'ADMIN':
            elections = Election.objects.all().order_by('-created_at')
            return Response(ElectionSerializer(elections, many=True).data)
            
        # Voter gets elections with voting eligibility metadata
        try:
            voter = user.voter_profile
        except Voter.DoesNotExist:
            return Response(
                {"error": "Voter profile not found."},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # Only show SCHEDULED, ACTIVE, COMPLETED to voters
        elections = Election.objects.exclude(status=Election.DRAFT).order_by('-start_date')
        
        data = []
        for election in elections:
            already_voted = VoteReceipt.objects.filter(voter=voter, election=election).exists()
            receipt = VoteReceipt.objects.filter(voter=voter, election=election).first()
            
            election_info = ElectionSerializer(election).data
            election_info['already_voted'] = already_voted
            election_info['receipt_number'] = receipt.receipt_number if receipt else None
            
            # Additional helper info for front-end
            election_info['is_verified_voter'] = voter.is_verified
            
            data.append(election_info)
            
        return Response(data)


class CandidateListCreateView(APIView):
    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdmin()]
        return [permissions.IsAuthenticated()]

    def get(self, request):
        user = request.user
        election_id = request.query_params.get('election_id')
        
        candidates = Candidate.objects.all().order_by('name')
        if election_id:
            candidates = candidates.filter(election_id=election_id)
            
        # If user is voter, filter candidates to match voter constituency and only show approved
        if user.role == 'VOTER':
            try:
                voter = user.voter_profile
                candidates = candidates.filter(
                    constituency=voter.constituency,
                    is_approved=True
                )
            except Voter.DoesNotExist:
                return Response(
                    {"error": "Voter profile required."},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        serializer = CandidateSerializer(candidates, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = CandidateCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        candidate = serializer.save()
        
        log_event(
            request.user,
            'CANDIDATE_CREATION',
            request,
            {'candidate_id': str(candidate.id), 'candidate_name': candidate.name}
        )
        return Response(CandidateSerializer(candidate).data, status=status.HTTP_201_CREATED)


class CandidateDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Candidate.objects.all()
    serializer_class = CandidateSerializer

    def get_permissions(self):
        if self.request.method == 'GET':
            return [permissions.IsAuthenticated()]
        return [IsAdmin()]


class ApproveCandidateView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            candidate = Candidate.objects.get(pk=pk)
        except Candidate.DoesNotExist:
            return Response(
                {"error": "Candidate not found."},
                status=status.HTTP_404_NOT_FOUND
            )
            
        candidate.is_approved = True
        candidate.save()
        
        log_event(
            request.user,
            'CANDIDATE_APPROVAL',
            request,
            {'candidate_id': str(candidate.id), 'candidate_name': candidate.name}
        )
        
        return Response(CandidateSerializer(candidate).data, status=status.HTTP_200_OK)


class VoteCastView(APIView):
    permission_classes = [IsVoter]

    def post(self, request):
        candidate_id = request.data.get('candidate_id')
        if not candidate_id:
            return Response(
                {"error": "Candidate ID is required."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        try:
            candidate = Candidate.objects.get(pk=candidate_id)
        except Candidate.DoesNotExist:
            return Response(
                {"error": "Candidate not found."},
                status=status.HTTP_404_NOT_FOUND
            )
            
        election = candidate.election
        
        # 1. Verify Voter Profile
        try:
            voter = request.user.voter_profile
        except Voter.DoesNotExist:
            return Response(
                {"error": "Voter profile not found."},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # 2. Check if verified
        if not voter.is_verified:
            return Response(
                {"error": "Your voter account has not been verified by an administrator. You cannot vote yet."},
                status=status.HTTP_403_FORBIDDEN
            )
            
        # 3. Check election status
        if election.status != Election.ACTIVE:
            return Response(
                {"error": f"Voting is not open for this election. Current status: {election.status}"},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 4. Check date bounds
        now = timezone.now()
        if now < election.start_date or now > election.end_date:
            return Response(
                {"error": "This election is either closed or has not started yet."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 5. Check candidate approval status
        if not candidate.is_approved:
            return Response(
                {"error": "This candidate is not approved for this election."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 6. Check voter constituency match
        if voter.constituency != candidate.constituency:
            return Response(
                {"error": "Constituency mismatch. You cannot vote for a candidate outside your constituency."},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        # 7. Atomic transaction voting submission
        try:
            with transaction.atomic():
                # Re-verify inside atomic transaction to lock record
                if VoteReceipt.objects.filter(voter=voter, election=election).exists():
                    return Response(
                        {"error": "Double voting detected. You have already cast a vote in this election."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                    
                # Generate unique receipt string
                salt = f"{voter.id}-{election.id}-{timezone.now().isoformat()}-{random.randint(100000, 999999)}"
                receipt_hash = hashlib.sha256(salt.encode()).hexdigest().upper()
                
                # Record receipt (linked to Voter + Election, unique together ensures single vote)
                receipt = VoteReceipt.objects.create(
                    election=election,
                    voter=voter,
                    receipt_number=receipt_hash
                )
                
                # Record Vote (linked ONLY to election, candidate and constituency - completely detached from voter)
                vote = Vote.objects.create(
                    election=election,
                    constituency=voter.constituency,
                    candidate=candidate
                )
                
                # Audit log - logs who voted in what election, but NEVER the candidate choice
                log_event(
                    request.user,
                    'VOTE_CAST',
                    request,
                    {
                        'election_id': str(election.id),
                        'receipt_number': receipt_hash,
                        'constituency_id': str(voter.constituency.id)
                    }
                )
        except IntegrityError:
            return Response(
                {"error": "Double voting detected. You have already cast a vote in this election."},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {"error": f"An error occurred while recording your vote: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )
            
        return Response({
            "message": "Your vote has been cast successfully.",
            "receipt_number": receipt_hash,
            "timestamp": receipt.timestamp,
            "election_title": election.title
        }, status=status.HTTP_201_CREATED)


class ElectionResultsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            election = Election.objects.get(pk=pk)
        except Election.DoesNotExist:
            return Response(
                {"error": "Election not found."},
                status=status.HTTP_404_NOT_FOUND
            )
            
        total_votes_cast = Vote.objects.filter(election=election).count()
        total_verified_voters = Voter.objects.filter(is_verified=True).count()
        
        turnout_percentage = 0.0
        if total_verified_voters > 0:
            turnout_percentage = round((total_votes_cast / total_verified_voters) * 100, 2)
            
        # Candidates standing
        candidates = Candidate.objects.filter(election=election)
        candidates_data = []
        for candidate in candidates:
            votes = Vote.objects.filter(candidate=candidate).count()
            candidates_data.append({
                "id": str(candidate.id),
                "name": candidate.name,
                "party_name": candidate.party_name,
                "constituency_name": candidate.constituency.name,
                "votes": votes
            })
            
        # Sort candidates by votes descending
        candidates_data.sort(key=lambda x: x['votes'], reverse=True)
        
        # Constituency-wise breakdown
        constituency_data = []
        for constituency in Constituency.objects.all():
            registered = Voter.objects.filter(constituency=constituency, is_verified=True).count()
            votes_cast = Vote.objects.filter(election=election, constituency=constituency).count()
            
            c_turnout = 0.0
            if registered > 0:
                c_turnout = round((votes_cast / registered) * 100, 2)
                
            constituency_data.append({
                "constituency_name": constituency.name,
                "registered_voters": registered,
                "votes_cast": votes_cast,
                "turnout_percentage": c_turnout
            })
            
        return Response({
            "election_id": str(election.id),
            "election_title": election.title,
            "status": election.status,
            "total_votes": total_votes_cast,
            "total_voters": total_verified_voters,
            "turnout_percentage": turnout_percentage,
            "candidates": candidates_data,
            "constituency_turnout": constituency_data
        })


class AdminSummaryView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        total_voters = Voter.objects.count()
        pending_voters = Voter.objects.filter(is_verified=False).count()
        active_elections = Election.objects.filter(status=Election.ACTIVE).count()
        pending_candidates = Candidate.objects.filter(is_approved=False).count()
        
        # Recent audit logs
        recent_logs = AuditLog.objects.all().order_by('-created_at')[:15]
        logs_data = []
        for log in recent_logs:
            logs_data.append({
                "id": str(log.id),
                "username": log.user.username if log.user else "Anonymous",
                "action": log.action,
                "ip_address": log.ip_address,
                "timestamp": log.created_at,
                "details": log.details
            })
            
        return Response({
            "total_voters": total_voters,
            "pending_voters": pending_voters,
            "active_elections": active_elections,
            "pending_candidates": pending_candidates,
            "recent_logs": logs_data
        })


class ExportElectionResultsExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            election = Election.objects.get(pk=pk)
        except Election.DoesNotExist:
            return Response(
                {"error": "Election not found."},
                status=status.HTTP_404_NOT_FOUND
            )
            
        from openpyxl import Workbook
        from django.http import HttpResponse
        
        wb = Workbook()
        
        # Sheet 1: Summary Info
        ws_summary = wb.active
        ws_summary.title = "Summary Metrics"
        
        ws_summary.append(["DigiVoting - Election Results Report"])
        ws_summary.append([])
        ws_summary.append(["Election ID", str(election.id)])
        ws_summary.append(["Election Title", election.title])
        ws_summary.append(["Current Status", election.status])
        ws_summary.append(["Start Date", election.start_date.strftime("%Y-%m-%d %H:%M:%S") if election.start_date else "N/A"])
        ws_summary.append(["End Date", election.end_date.strftime("%Y-%m-%d %H:%M:%S") if election.end_date else "N/A"])
        
        total_votes = Vote.objects.filter(election=election).count()
        total_voters = Voter.objects.filter(is_verified=True).count()
        turnout = (total_votes / total_voters * 100) if total_voters > 0 else 0.0
        
        ws_summary.append(["Total Verified Voters", total_voters])
        ws_summary.append(["Total Votes Cast", total_votes])
        ws_summary.append(["Turnout Rate", f"{turnout:.2f}%"])
        
        # Sheet 2: Candidate Standings
        ws_candidates = wb.create_sheet(title="Candidate Standings")
        ws_candidates.append(["Candidate ID", "Candidate Name", "Political Party", "Constituency", "Votes Received"])
        candidates = Candidate.objects.filter(election=election)
        for c in candidates:
            votes = Vote.objects.filter(candidate=c).count()
            ws_candidates.append([str(c.id), c.name, c.party_name, c.constituency.name, votes])
            
        # Sheet 3: Constituency Turnout
        ws_constituencies = wb.create_sheet(title="Constituency Turnout")
        ws_constituencies.append(["Constituency Name", "Registered Voters", "Votes Cast", "Turnout Percentage"])
        for con in Constituency.objects.all():
            reg = Voter.objects.filter(constituency=con, is_verified=True).count()
            cast = Vote.objects.filter(election=election, constituency=con).count()
            c_pct = (cast / reg * 100) if reg > 0 else 0.0
            ws_constituencies.append([con.name, reg, cast, f"{c_pct:.2f}%"])
            
        # Adjust column widths for basic readability
        for ws in [ws_summary, ws_candidates, ws_constituencies]:
            for col in ws.columns:
                max_len = max(len(str(val or '')) for val in [cell.value for cell in col])
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="election_report_{election.title.replace(" ", "_")}.xlsx"'
        wb.save(response)
        return response

